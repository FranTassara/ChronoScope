"""
generate_user_template.py
=========================

Generates a clean, user-friendly data template for ChronoScope.

Two files are produced (same folder as this script):
  user_data_template.csv   — universal CSV, opens in Excel / LibreOffice / any editor
  user_data_template.xlsx  — Excel with formatted headers and an Instructions sheet

HOW TO USE THE TEMPLATE
-----------------------
1. Open either file.
2. Replace the example values with your own measurements.
3. Rename "gene_expression" and "locomotor_activity" to your variable names
   (e.g. "body_temperature", "cortisol", "Per2_mRNA").
4. Add or remove variable columns as needed — ChronoScope auto-detects them.
5. Save as CSV or Excel and load in ChronoScope.

Usage
-----
  python examples/generate_user_template.py
"""

import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from pathlib import Path

# ---------------------------------------------------------------------------
# Template parameters
# ---------------------------------------------------------------------------
TIMEPOINTS   = [0, 4, 8, 12, 16, 20]          # abbreviated 24-h cycle (hours)
CONDITIONS   = ["control", "treatment"]
REPLICATES   = [1, 2, 3]

# Deterministic seed so every run produces the same values
RNG = np.random.default_rng(seed=0)

PERIOD_RAD = 2 * np.pi / 24                     # 24-h period


def _cosinor(t, mesor, amp, acrophase, noise_sd):
    return mesor + amp * np.cos(PERIOD_RAD * t - acrophase) + RNG.normal(0, noise_sd)


# Variable parameters: (mesor, amplitude, acrophase_control, noise_sd)
VARIABLES = {
    "gene_expression":     (10.0, 2.5, np.pi / 4,        0.4),   # peaks ~ZT6
    "locomotor_activity":  (50.0, 15.0, np.pi / 6,       2.0),   # peaks ~ZT4
}
TREATMENT_PHASE_SHIFT = np.pi / 4   # +6 h shift for treatment group


# ---------------------------------------------------------------------------
# Build DataFrame
# ---------------------------------------------------------------------------
def build_template_df() -> pd.DataFrame:
    rows = []
    for condition in CONDITIONS:
        phase_extra = TREATMENT_PHASE_SHIFT if condition == "treatment" else 0.0
        for rep in REPLICATES:
            for t in TIMEPOINTS:
                row = {"time": t, "condition": condition, "replicate": rep}
                for var_name, (mesor, amp, acrophase, noise) in VARIABLES.items():
                    row[var_name] = round(
                        _cosinor(t, mesor, amp, acrophase + phase_extra, noise), 3
                    )
                rows.append(row)

    df = pd.DataFrame(rows)
    df = df.sort_values(["time", "condition", "replicate"]).reset_index(drop=True)
    return df


# ---------------------------------------------------------------------------
# Write CSV
# ---------------------------------------------------------------------------
def write_csv(df: pd.DataFrame, path: Path) -> None:
    df.to_csv(path, index=False)
    print(f"CSV  saved: {path}")


# ---------------------------------------------------------------------------
# Write Excel (formatted)
# ---------------------------------------------------------------------------
# Color palette
_BLUE_HEADER  = "1F4E79"   # structural columns (time, condition, replicate)
_GREEN_HEADER = "375623"   # variable columns
_LIGHT_BLUE   = "D6E4F0"   # even-row highlight for structural cols
_LIGHT_GREEN  = "EBF1DE"   # even-row highlight for variable cols
_WHITE        = "FFFFFF"
_GOLD         = "FFC000"
_DARK_GRAY    = "404040"


def _header_fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _thin_border() -> Border:
    s = Side(style="thin", color="AAAAAA")
    return Border(left=s, right=s, top=s, bottom=s)


def _write_data_sheet(ws, df: pd.DataFrame) -> None:
    """Write the Data sheet with formatted headers and banded rows."""
    structural_cols = {"time", "condition", "replicate"}
    headers = df.columns.tolist()

    # --- Header row ---
    for col_idx, col_name in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        is_structural = col_name in structural_cols
        cell.fill  = _header_fill(_BLUE_HEADER if is_structural else _GREEN_HEADER)
        cell.font  = Font(bold=True, color=_WHITE, size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _thin_border()

    # --- Data rows ---
    for row_idx, (_, row) in enumerate(df.iterrows(), start=2):
        is_even = (row_idx % 2 == 0)
        for col_idx, col_name in enumerate(headers, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row[col_name])
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = _thin_border()
            if is_even:
                is_structural = col_name in structural_cols
                cell.fill = _header_fill(_LIGHT_BLUE if is_structural else _LIGHT_GREEN)

    # --- Column widths ---
    col_widths = {"time": 8, "condition": 14, "replicate": 12}
    default_var_width = 20
    for col_idx, col_name in enumerate(headers, start=1):
        width = col_widths.get(col_name, default_var_width)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # --- Row height ---
    ws.row_dimensions[1].height = 28

    # --- Freeze header row ---
    ws.freeze_panes = "A2"

    # --- Auto-filter ---
    ws.auto_filter.ref = ws.dimensions


def _write_instructions_sheet(ws, variable_names: list[str]) -> None:
    """Write the Instructions sheet."""

    def _heading(row, text):
        cell = ws.cell(row=row, column=1, value=text)
        cell.font = Font(bold=True, size=13, color=_BLUE_HEADER)
        ws.row_dimensions[row].height = 22
        return row + 1

    def _subheading(row, text):
        cell = ws.cell(row=row, column=1, value=text)
        cell.font = Font(bold=True, size=11, color=_DARK_GRAY)
        ws.row_dimensions[row].height = 18
        return row + 1

    def _body(row, text, indent=0):
        cell = ws.cell(row=row, column=1, value=("    " * indent) + text)
        cell.font = Font(size=11)
        cell.alignment = Alignment(wrap_text=True)
        ws.row_dimensions[row].height = 15
        return row + 1

    def _table_row(row, col_name, role, description, color=None):
        for c, val in enumerate([col_name, role, description], start=1):
            cell = ws.cell(row=row, column=c, value=val)
            cell.border = _thin_border()
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if color:
                cell.fill = _header_fill(color)
                cell.font = Font(bold=True, color=_WHITE, size=10)
            else:
                cell.font = Font(size=10)
        ws.row_dimensions[row].height = 18
        return row + 1

    # ---- Content ----
    r = 1

    ws.cell(row=r, column=1, value="ChronoScope — Data Template Guide").font = Font(
        bold=True, size=16, color=_BLUE_HEADER
    )
    ws.row_dimensions[r].height = 28
    r += 2

    r = _heading(r, "Required columns")
    r = _table_row(r, "Column", "Role", "Description", color=_BLUE_HEADER)
    r = _table_row(r, "time",      "Structural",  "Time in hours (e.g. 0, 4, 8, 12, 16, 20, 24). Must be numeric.")
    r = _table_row(r, "condition", "Structural",  'Group label (e.g. "control", "WT", "KO"). Can be any text.')
    r = _table_row(r, "replicate", "Optional",    "Biological replicate number (1, 2, 3…). Optional but recommended.")
    r += 1

    r = _heading(r, "Variable columns (rename to match your experiment)")
    r = _table_row(r, "Column", "Role", "How to use", color=_GREEN_HEADER)
    for var in variable_names:
        r = _table_row(r, var, "Variable", "Rename this to your measurement (e.g. 'Per2_mRNA', 'cortisol', 'body_temp').")
    r = _table_row(r, "…", "Variable", "Add as many variable columns as needed. ChronoScope auto-detects all numeric columns.")
    r += 1

    r = _heading(r, "Key rules")
    rules = [
        "• Time must be numeric (hours). Typical range: 0–48 h for a 48-h recording.",
        "• Each row is one measurement from one biological replicate at one timepoint.",
        "• Multiple rows with the same (time, condition) are treated as replicates.",
        "• Missing values are allowed — leave the cell empty; ChronoScope skips them.",
        "• Do NOT add extra header rows, merged cells, or comment rows above the header.",
        "• Save as .csv (comma-separated) or .xlsx — both are supported.",
    ]
    for rule in rules:
        r = _body(r, rule)
    r += 1

    r = _heading(r, "Example study designs")
    designs = [
        ("Independent (cross-sectional)", "Different animals at each timepoint. No subject column needed."),
        ("Dependent (longitudinal)",      "Same animals measured repeatedly. Add a 'subject' column (e.g. 'mouse1')."),
        ("Single condition",              "Only one condition group (omit condition column or use a single value)."),
        ("Multiple variables",            "Add columns: 'Per2', 'Bmal1', 'Rev-erb'. All are analysed together."),
    ]
    r = _table_row(r, "Design", "When to use", "Notes", color=_DARK_GRAY)
    for name, desc in designs:
        r = _table_row(r, name, "", desc)
    r += 1

    r = _heading(r, "Quick-start checklist")
    checklist = [
        "[ ] Rename 'gene_expression' and 'locomotor_activity' to your variable names",
        "[ ] Replace example values with your measurements",
        "[ ] Verify: time column is numeric (no 'ZT' prefix — use 0, not 'ZT0')",
        "[ ] Verify: at least 5–6 distinct timepoints per condition",
        "[ ] Verify: at least 3 replicates per timepoint for population analysis",
        "[ ] Save and load in ChronoScope: File > Load Data",
    ]
    for item in checklist:
        r = _body(r, item)

    # Column widths
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 65
    ws.sheet_view.showGridLines = False


def write_excel(df: pd.DataFrame, path: Path) -> None:
    wb = openpyxl.Workbook()

    # Sheet 1: Data
    ws_data = wb.active
    ws_data.title = "Data"
    _write_data_sheet(ws_data, df)

    # Sheet 2: Instructions
    ws_instr = wb.create_sheet("Instructions")
    _write_instructions_sheet(ws_instr, variable_names=list(VARIABLES.keys()))

    # Make Instructions the active/visible sheet on open
    wb.active = ws_instr

    wb.save(path)
    print(f"Excel saved: {path}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
if __name__ == "__main__":
    out_dir = Path(__file__).parent
    df = build_template_df()

    write_csv(df,   out_dir / "user_data_template.csv")
    write_excel(df, out_dir / "user_data_template.xlsx")

    print(f"\nShape: {df.shape}  |  Columns: {df.columns.tolist()}")
    print("\nPreview (first 8 rows):")
    print(df.head(8).to_string(index=False))
    print("\nDone. Load either file in ChronoScope: File > Load Data.")
